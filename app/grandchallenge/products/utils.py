from django.core.files.images import ImageFile
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from grandchallenge.products.models import (
    Company,
    Product,
    ProductImage,
    Status,
)

STATUS_MAPPING = {
    "Certified": Status.CERTIFIED,
    "No or not yet": Status.NO,
    "Not applicable": Status.NA,
    "510(k) cleared": Status.CLEARED,
    "De novo 510(k) cleared": Status.DE_NOVO_CLEARED,
    "PMA approved": Status.PMA_APPROVED,
    "Yes": Status.YES,
    "No": Status.NO,
}

CE_MAPPING = {
    "Medical Devices Directive (MDD)": "MDD",
    "Medical Device Regulation (MDR)": "MDR",
    "Medical Devices Directive (MDD), Medical Device Regulation (MDR)": "MDD/MDR",
}


class Row:
    def __init__(self, *, row, locs):
        self._row = row
        self._locs = locs

    def __getitem__(self, item):
        value = self._row[self._locs[item]]
        if value is None:
            return ""
        else:
            return value


class Sheet:
    def __init__(self, *, filename):
        self._workbook = load_workbook(filename=filename)
        self._sheet = self._workbook.active
        self._locs = {cell.value: cell.column - 1 for cell in self._sheet["1"]}

    def iter_rows(self):
        for row in self._sheet.iter_rows(min_row=2, values_only=True):
            yield Row(row=row, locs=self._locs)


def import_data(*, products_path, companies_path, images_path):
    companies = Sheet(filename=companies_path)
    products = Sheet(filename=products_path)

    with transaction.atomic():
        # Delete all existing entries and recreate them
        Company.objects.all().delete()
        Product.objects.all().delete()
        ProductImage.objects.all().delete()
        pk = "Company name"

        for company_row in companies.iter_rows():
            if not company_row[pk]:
                break

            company = _create_company(row=company_row, images_path=images_path)

            for product_row in products.iter_rows():
                if not product_row[pk]:
                    break

                if product_row[pk] != company_row[pk]:
                    continue

                product = _create_product(row=product_row, company=company)
                _create_product_images(
                    product=product,
                    row=product_row,
                    images_path=images_path,
                )


def _split(string, max_char):
    if len(string) > max_char:
        short_string = string[:max_char].rsplit(" ", 1)[0] + " ..."
    else:
        return string
    return short_string


def _create_company(*, row, images_path):
    company = Company()
    company.company_name = row["Company name"]
    company.modified = row["Timestamp"]
    company.website = row["Company website url"]
    company.founded = row["Founded"]
    company.hq = row["Head office"]
    company.email = row["Email address (public)"]
    company.description = row["Company description"]
    company.description_short = _split(row["Company description"], 200)
    slug = slugify(row["Company name"])
    company.slug = slug
    img_file = images_path.glob(f"**/logo/{slug}.*")

    for file in img_file:
        company.logo = ImageFile(open(file, "rb"))

    company.save()

    return company


def _create_product(*, row, company):
    product = Product()
    product.product_name = row["Product name"]
    product.company = company
    product.modified = row["Timestamp"]
    product.slug = slugify(row["Short name"])
    product.description = row["Product description"]
    product.description_short = _split(row["Product description"], 200)
    product.modality = row["Modality"]
    product.subspeciality = row["Subspeciality"]
    product.input_data = row["Input data"]
    product.file_format_input = row["File format of input data"]
    product.output_data = row["Output data"]
    product.file_format_output = row["File format of output data"]
    product.key_features = row["Key-feature(s)"]
    product.ce_status = STATUS_MAPPING.get(row["CE-certified"], Status.UNKNOWN)
    product.ce_under = CE_MAPPING.get(row["CE-certified under"], "")
    product.ce_class = row["If CE-certified, what class"]
    product.fda_status = STATUS_MAPPING.get(
        row["FDA approval/clearance"], Status.UNKNOWN
    )
    product.fda_class = row["If FDA approval/clearance, what class"]
    product.verified = STATUS_MAPPING.get(row["Verified"], Status.UNKNOWN)
    product.ce_verified = STATUS_MAPPING.get(
        row["CE verified"], Status.UNKNOWN
    )
    product.integration = row["Integration"]
    product.deployment = row["Deployment"]
    product.process_time = row["Algorithm processing time per study"]
    product.trigger = row["Trigger for the analysis of data"]
    product.market_since = str(row["Product on the market since"])
    product.countries = str(row["Number of countries present"])
    product.diseases = row["Disease(s) targeted"]
    product.population = row["Population on which analysis is applied"]
    product.distribution = str(
        row["Distribution platforms/marketplaces availability"]
    )
    product.intended_use_ce = row[
        "If CE-certified, provide intended use according to the certification"
    ]
    product.intended_use_ce_public = row["Intended purpose public"]
    product.intended_use_fda = row[
        "If FDA approval/clearance, provide the intended use according to the approval"
    ]
    product.institutes_research = str(
        row["Number of institutes using the product for research"]
    )
    product.institutes_clinic = str(
        row["Number of institutes using the product in clinical practice"]
    )
    product.pricing_model = row["Pricing model"]
    product.pricing_basis = row["Pricing model based on"]
    product.tech_peer_papers = row[
        "Name peer reviewed papers that describe the performance of the software as is commercially available."
    ]
    product.tech_other_papers = row[
        "Name other (white)papers that describe the performance of the software as is commercially available."
    ]
    product.all_other_papers = row[
        "Name other relevant (white)papers regarding the performance or implementation of the software not mentioned above."
    ]

    product.save()

    return product


def _create_product_images(*, row, product, images_path):
    images = []
    img_files = images_path.glob(
        "**/product_images/{}*".format(row["Short name"])
    )
    product.images.all().delete()
    for file in img_files:
        image = ProductImage(img=ImageFile(open(file, "rb")))
        image.save()
        images.append(image)
    product.images.set(images)
